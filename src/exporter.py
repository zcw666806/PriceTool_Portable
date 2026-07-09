from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .database import query_prices, query_table
from .models import EXPORT_PRICE_COLUMNS


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")


def export_workbook(conn, output_folder: str | Path, filters: dict | None = None) -> Path:
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)
    output_path = output_folder / f"整理后的价格汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "All Prices"
    prices = query_prices(conn, filters=filters or {}, limit=None)
    _write_price_sheet(ws, prices)

    review_ws = wb.create_sheet("Review Needed")
    review_rows = [row for row in prices if row.get("needs_review")]
    _write_review_sheet(review_ws, review_rows)

    raw_ws = wb.create_sheet("PDF Extract Raw")
    _write_generic_sheet(raw_ws, query_table(conn, "pdf_extract_raw", limit=None))

    excel_raw_ws = wb.create_sheet("Excel Imported Raw")
    _write_generic_sheet(excel_raw_ws, query_table(conn, "excel_imported_raw", limit=None))

    products_ws = wb.create_sheet("Products")
    _write_generic_sheet(products_ws, query_table(conn, "products", limit=None))

    colours_ws = wb.create_sheet("Colour Reference")
    _write_generic_sheet(colours_ws, query_table(conn, "colour_reference", limit=None))

    log_ws = wb.create_sheet("Import Log")
    _write_generic_sheet(log_ws, query_table(conn, "source_files", limit=None))

    for sheet in wb.worksheets:
        _finish_sheet(sheet)

    wb.save(output_path)
    return output_path


def _write_price_sheet(ws, rows: list[dict]) -> None:
    headers = [header for header, _ in EXPORT_PRICE_COLUMNS]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(field) for _, field in EXPORT_PRICE_COLUMNS])
        if row.get("needs_review"):
            for cell in ws[ws.max_row]:
                cell.fill = REVIEW_FILL


def _write_review_sheet(ws, rows: list[dict]) -> None:
    headers = [
        "Review Reason",
        "Product Code",
        "Product Name",
        "Cover Range",
        "Size",
        "Raw Price",
        "Clean Price",
        "Source File",
        "Page",
        "Suggested Action",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.get("review_reason"),
                row.get("product_code"),
                row.get("product_name"),
                row.get("cover_range"),
                row.get("size"),
                row.get("raw_price"),
                row.get("price"),
                row.get("source_file"),
                row.get("page"),
                "人工核对来源文件",
            ]
        )


def _write_generic_sheet(ws, rows: list[dict]) -> None:
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _finish_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 45))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
