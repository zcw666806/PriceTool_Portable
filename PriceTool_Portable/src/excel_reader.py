from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .normalizer import build_lookup_key, is_probable_size, normalize_key, normalize_size, normalize_space
from .price_cleaner import clean_price


LONG_FIELD_MAP = {
    "PRODUCT CODE": "product_code",
    "PRODUCT NAME": "product_name",
    "SECTION": "section",
    "TIER": "tier",
    "COVER RANGE": "cover_range",
    "COVER TYPE": "cover_type",
    "SIZE": "size_raw",
    "PRICE (USD)": "price",
    "PRICE": "price",
    "LOOKUP KEY": "lookup_key",
    "SUPPLIER": "supplier",
    "VERSION": "version",
    "CURRENCY": "currency",
}


def detect_workbook_profile(path: str | Path) -> dict:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    names = set(workbook.sheetnames)
    lower_name = path.name.lower()
    if "All Prices" in names:
        if "ez" in lower_name or "living" in lower_name:
            return {"profile_code": "EZ_ALL_PRICES", "workbook_type": "long_table", "source_type": "EZ_EXCEL", "price_basis": "EZ"}
        return {"profile_code": "FOB_ALL_PRICES", "workbook_type": "long_table", "source_type": "FOB_EXCEL", "price_basis": "FOB"}
    if "sterling" in lower_name or "cif" in lower_name:
        return {"profile_code": "STERLING_WIDE", "workbook_type": "product_sheet_wide", "source_type": "CIF_EXCEL", "price_basis": "CIF"}
    if "fv" in lower_name:
        return {"profile_code": "FV_WIDE", "workbook_type": "matrix_wide", "source_type": "FV_EXCEL", "price_basis": "FV"}
    return {"profile_code": "UNKNOWN_EXCEL", "workbook_type": "unknown", "source_type": "EXCEL", "price_basis": None}


def import_excel_workbook(path: str | Path, import_batch_id: str | None = None) -> dict:
    path = Path(path)
    profile = detect_workbook_profile(path)
    data_wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    formula_wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)

    result = {
        "profile": profile,
        "price_rows": [],
        "products": [],
        "colour_reference": [],
        "raw_rows": [],
        "sheet_count": len(data_wb.sheetnames),
    }

    if profile["workbook_type"] == "long_table":
        result["price_rows"].extend(_read_long_prices(path, data_wb, profile, import_batch_id))
        result["products"].extend(_read_products(path, data_wb))
        result["colour_reference"].extend(_read_colour_reference(path, data_wb))
    elif profile["workbook_type"] in {"product_sheet_wide", "matrix_wide"}:
        result["price_rows"].extend(_read_wide_prices(path, data_wb, formula_wb, profile, import_batch_id))

    result["raw_rows"].extend(_read_excel_raw(path, data_wb, formula_wb))
    return result


def _read_long_prices(path: Path, wb, profile: dict, import_batch_id: str | None) -> list[dict]:
    if "All Prices" not in wb.sheetnames:
        return []
    ws = wb["All Prices"]
    header_row, headers = _find_header(ws, min_matches=4)
    if not header_row:
        return []
    rows = []
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = dict(zip(headers, excel_row))
        if not any(values.values()):
            continue
        mapped = {}
        for header, value in values.items():
            field = LONG_FIELD_MAP.get(normalize_key(header))
            if field:
                mapped[field] = value
        raw_price = mapped.get("price")
        cleaned = clean_price(raw_price)
        size = normalize_size(mapped.get("size_raw") or mapped.get("size"))
        product_code = normalize_space(mapped.get("product_code"))
        product_name = normalize_space(mapped.get("product_name"))
        review_reasons = [
            reason
            for reason in (
                size.get("review_reason") if size.get("needs_review") else None,
                cleaned.get("review_reason") if cleaned.get("needs_review") else None,
            )
            if reason
        ]
        rows.append(
            {
                "source_file": path.name,
                "source_path": str(path),
                "source_type": profile["source_type"],
                "source_role": "OPTIONAL_REFERENCE",
                "source_profile": profile["profile_code"],
                "import_batch_id": import_batch_id,
                "supplier": normalize_space(mapped.get("supplier")),
                "price_basis": profile.get("price_basis"),
                "currency": normalize_space(mapped.get("currency")) or "USD",
                "version": normalize_space(mapped.get("version")),
                "effective_date": None,
                "product_code": product_code,
                "product_name": product_name,
                "collection": None,
                "section": normalize_space(mapped.get("section")),
                "tier": normalize_space(mapped.get("tier")),
                "cover_range": normalize_space(mapped.get("cover_range")),
                "cover_type": normalize_space(mapped.get("cover_type")),
                "size": size.get("size"),
                "size_raw": size.get("size_raw"),
                "price": cleaned.get("price"),
                "raw_price": cleaned.get("raw_price"),
                "formula": None,
                "page": None,
                "table_index": None,
                "row_index": None,
                "confidence": 0.95 if not review_reasons else 0.7,
                "needs_review": bool(review_reasons),
                "review_reason": "; ".join(review_reasons),
                "lookup_key": normalize_space(mapped.get("lookup_key"))
                or build_lookup_key(product_code, product_name, mapped.get("tier"), mapped.get("cover_range"), size.get("size")),
            }
        )
    return rows


def _read_products(path: Path, wb) -> list[dict]:
    if "Products" not in wb.sheetnames:
        return []
    ws = wb["Products"]
    header_row, headers = _find_header(ws, min_matches=2)
    if not header_row:
        return []
    rows = []
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = dict(zip(headers, excel_row))
        if not any(values.values()):
            continue
        rows.append(
            {
                "product_code": values.get("Product Code") or values.get("Code"),
                "product_name": values.get("Product Name") or values.get("Product"),
                "supplier": values.get("Supplier"),
                "alias_name": values.get("Alias Name") or values.get("Alias"),
                "source_file": path.name,
                "active": 1,
            }
        )
    return rows


def _read_colour_reference(path: Path, wb) -> list[dict]:
    sheet_name = "Colour Reference" if "Colour Reference" in wb.sheetnames else None
    if not sheet_name:
        return []
    ws = wb[sheet_name]
    header_row, headers = _find_header(ws, min_matches=2)
    if not header_row:
        return []
    rows = []
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = dict(zip(headers, excel_row))
        if not any(values.values()):
            continue
        rows.append(
            {
                "colour_code": values.get("Colour Code") or values.get("Code"),
                "colour_name": values.get("Colour Name") or values.get("Colour"),
                "category": values.get("Category"),
                "cover_range": values.get("Cover Range"),
                "price_tier": values.get("Price Tier") or values.get("Tier"),
                "cover_type": values.get("Cover Type"),
                "notes": values.get("Notes"),
                "source_file": path.name,
            }
        )
    return rows


def _read_wide_prices(path: Path, data_wb, formula_wb, profile: dict, import_batch_id: str | None) -> list[dict]:
    rows = []
    for sheet_name in data_wb.sheetnames:
        if sheet_name.lower().startswith(("index", "contents", "summary")):
            continue
        ws = data_wb[sheet_name]
        formula_ws = formula_wb[sheet_name]
        header_row, headers = _find_wide_header(ws)
        if not header_row:
            continue
        meta_columns, size_columns = _classify_wide_columns(headers)
        product_code, product_name = _parse_product_from_sheet(sheet_name)
        carry = {"section": None, "tier": None, "cover_range": None, "cover_type": None}
        last_row = min(ws.max_row or header_row, header_row + 800)
        last_col = min(len(headers), 80)
        for row_number in range(header_row + 1, last_row + 1):
            row_values = [ws.cell(row=row_number, column=col).value for col in range(1, last_col + 1)]
            if not any(row_values):
                continue
            for field, col_idx in meta_columns.items():
                value = normalize_space(row_values[col_idx] if col_idx < len(row_values) else "")
                if value:
                    carry[field] = value
            for col_idx, raw_size in size_columns.items():
                cell = ws.cell(row=row_number, column=col_idx + 1)
                raw_price = cell.value
                cleaned = clean_price(raw_price)
                if cleaned["price"] is None and not normalize_space(raw_price):
                    continue
                formula_value = formula_ws.cell(row=row_number, column=col_idx + 1).value
                formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
                size = normalize_size(raw_size)
                review_reasons = [
                    reason
                    for reason in (
                        "Product code missing in sheet name" if not product_code else None,
                        size.get("review_reason") if size.get("needs_review") else None,
                        cleaned.get("review_reason") if cleaned.get("needs_review") else None,
                    )
                    if reason
                ]
                rows.append(
                    {
                        "source_file": path.name,
                        "source_path": str(path),
                        "source_type": profile["source_type"],
                        "source_role": "OPTIONAL_REFERENCE",
                        "source_profile": profile["profile_code"],
                        "import_batch_id": import_batch_id,
                        "supplier": None,
                        "price_basis": profile.get("price_basis"),
                        "currency": "GBP" if profile.get("price_basis") in {"CIF", "FV"} else "USD",
                        "version": None,
                        "effective_date": sheet_name if profile["profile_code"] == "FV_WIDE" else None,
                        "product_code": product_code,
                        "product_name": product_name,
                        "collection": None,
                        "section": carry.get("section"),
                        "tier": carry.get("tier"),
                        "cover_range": carry.get("cover_range"),
                        "cover_type": carry.get("cover_type"),
                        "size": size.get("size"),
                        "size_raw": size.get("size_raw"),
                        "price": cleaned.get("price"),
                        "raw_price": cleaned.get("raw_price"),
                        "formula": formula,
                        "page": None,
                        "table_index": None,
                        "row_index": row_number,
                        "confidence": 0.9 if not review_reasons else 0.65,
                        "needs_review": bool(review_reasons),
                        "review_reason": "; ".join(review_reasons),
                        "lookup_key": build_lookup_key(product_code, product_name, carry.get("tier"), carry.get("cover_range"), size.get("size")),
                    }
                )
    return rows


def _find_header(ws, min_matches: int = 3) -> tuple[int | None, list[str]]:
    for row_number in range(1, min(ws.max_row, 25) + 1):
        values = [normalize_space(cell.value) for cell in ws[row_number]]
        matches = sum(1 for value in values if normalize_key(value) in LONG_FIELD_MAP)
        if matches >= min_matches:
            return row_number, values
    return None, []


def _find_wide_header(ws) -> tuple[int | None, list[str]]:
    for row_number in range(1, min(ws.max_row, 30) + 1):
        values = [normalize_space(cell.value) for cell in ws[row_number]]
        size_count = sum(1 for value in values if is_probable_size(value))
        if size_count >= 2:
            return row_number, values
    return None, []


def _classify_wide_columns(headers: list[str]) -> tuple[dict[str, int], dict[int, str]]:
    meta_columns = {}
    size_columns = {}
    for idx, header in enumerate(headers):
        key = normalize_key(header)
        if key in {"FABRIC/LEATHER", "SECTION"}:
            meta_columns["section"] = idx
        elif key in {"TIER", "GRADE"}:
            meta_columns["tier"] = idx
        elif key in {"COVER", "COVER RANGE", "RANGE"}:
            meta_columns["cover_range"] = idx
        elif key in {"TYPE", "COVER TYPE"}:
            meta_columns["cover_type"] = idx
        elif is_probable_size(header):
            size_columns[idx] = header
    return meta_columns, size_columns


def _parse_product_from_sheet(sheet_name: str) -> tuple[str | None, str]:
    match = re.match(r"^\s*(?P<code>\d{3,6})\s+(?P<name>.+)$", sheet_name)
    if match:
        return match.group("code"), match.group("name").strip()
    return None, sheet_name.strip()


def _read_excel_raw(path: Path, data_wb, formula_wb, max_cells: int = 50000) -> list[dict]:
    rows = []
    for sheet_name in data_wb.sheetnames:
        ws = data_wb[sheet_name]
        formula_ws = formula_wb[sheet_name]
        max_row = min(ws.max_row or 1, 1500)
        max_col = min(ws.max_column or 1, 80)
        data_rows = ws.iter_rows(max_row=max_row, max_col=max_col)
        formula_rows = formula_ws.iter_rows(max_row=max_row, max_col=max_col)
        for row_number, (data_row, formula_row) in enumerate(zip(data_rows, formula_rows), start=1):
            for column_number, (cell, formula_cell) in enumerate(zip(data_row, formula_row), start=1):
                value = cell.value
                formula_value = formula_cell.value
                formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
                if value is None and formula is None:
                    continue
                rows.append(
                    {
                        "source_file": path.name,
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                        "column_number": column_number,
                        "value": None if value is None else str(value),
                        "formula": formula,
                    }
                )
                if len(rows) >= max_cells:
                    return rows
    return rows
